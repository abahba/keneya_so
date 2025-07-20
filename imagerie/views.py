from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import ExamenImagerie, HistoriqueImagerie
from .forms import ExamenImagerieForm
from django.http import HttpResponse
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.utils import get_column_letter

@login_required
def liste_examens_imagerie(request):
    """Liste tous les examens d'imagerie avec tri par date décroissante"""
    examens = ExamenImagerie.objects.select_related('patient', 'type_imagerie', 'prescripteur')\
                                   .prefetch_related('historiques')\
                                   .order_by('-date_prescription')
    return render(request, 'imagerie/liste_examens.html', {'examens': examens})

@login_required
def ajouter_examen_imagerie(request):
    """Ajoute un nouvel examen d'imagerie"""
    if request.method == 'POST':
        form = ExamenImagerieForm(request.POST, request.FILES)
        if form.is_valid():
            examen = form.save(commit=False)
            examen.utilisateur = request.user
            examen.save()
            
            HistoriqueImagerie.objects.create(
                examen=examen,
                utilisateur=request.user,
                action='ajout',
                commentaire=f'Ajout d\'un nouvel examen: {examen.type_imagerie}'
            )
            messages.success(request, 'Examen ajouté avec succès!')
            return redirect('liste_examens_imagerie')
    else:
        form = ExamenImagerieForm(initial={'prescripteur': request.user})
    
    return render(request, 'imagerie/ajouter_examen.html', {'form': form})

@login_required
def modifier_examen_imagerie(request, examen_id):
    """Modifie un examen existant"""
    examen = get_object_or_404(ExamenImagerie, id=examen_id)
    
    if request.method == 'POST':
        form = ExamenImagerieForm(request.POST, request.FILES, instance=examen)
        if form.is_valid():
            examen = form.save()
            
            HistoriqueImagerie.objects.create(
                examen=examen,
                utilisateur=request.user,
                action='modification',
                commentaire=f'Modification de l\'examen {examen.id}'
            )
            messages.success(request, 'Examen modifié avec succès!')
            return redirect('liste_examens_imagerie')
    else:
        form = ExamenImagerieForm(instance=examen)
    
    return render(request, 'imagerie/modifier_examen.html', {
        'form': form,
        'examen': examen
    })

@login_required
def supprimer_examen_imagerie(request, examen_id):
    """Supprime un examen après confirmation"""
    examen = get_object_or_404(ExamenImagerie, id=examen_id)
    
    if request.method == 'POST':
        HistoriqueImagerie.objects.create(
            examen=examen,
            utilisateur=request.user,
            action='suppression',
            commentaire=f'Suppression de l\'examen {examen.id}'
        )
        examen.delete()
        messages.success(request, 'Examen supprimé avec succès!')
        return redirect('liste_examens_imagerie')
    
    return render(request, 'imagerie/supprimer_examen.html', {'examen': examen})

@login_required
def historique_examen_imagerie(request, examen_id):
    examen = get_object_or_404(ExamenImagerie, pk=examen_id)
    historiques = examen.historiques.all()

    # Récupération des filtres depuis le formulaire
    action = request.GET.get('action')
    date = request.GET.get('date')

    # Filtrage si les champs sont remplis
    if action:
        historiques = historiques.filter(action=action)
    if date:
        historiques = historiques.filter(date_action__date=date)

    context = {
        'examen': examen,
        'historiques': historiques,
        'action_filter': action or '',
        'date_filter': date or '',
    }
    return render(request, 'imagerie/historique_examen.html', context)


@login_required
def export_historique_imagerie_pdf(request, examen_id):
    examen = get_object_or_404(ExamenImagerie, pk=examen_id)
    historiques = examen.historiques.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="historique_imagerie_{examen.id}.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 800, f"Historique de l'examen {examen} - {examen.patient.nom}")

    y = 780
    for h in historiques:
        ligne = f"{h.date_action.strftime('%d/%m/%Y %H:%M')} - {h.get_action_display()} - {h.utilisateur} - {h.commentaire}"
        p.drawString(50, y, ligne)
        y -= 20
        if y < 50:
            p.showPage()
            y = 800

    p.showPage()
    p.save()
    return response

@login_required
def export_historique_imagerie_excel(request, examen_id):
    examen = get_object_or_404(ExamenImagerie, pk=examen_id)
    historiques = examen.historiques.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Historique Imagerie"

    headers = ['Date', 'Action', 'Utilisateur', 'Commentaire']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header

    for row_num, historique in enumerate(historiques, start=2):
        sheet[f'A{row_num}'] = historique.date_action.strftime('%d/%m/%Y %H:%M')
        sheet[f'B{row_num}'] = historique.get_action_display()
        sheet[f'C{row_num}'] = str(historique.utilisateur)
        sheet[f'D{row_num}'] = historique.commentaire

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=historique_imagerie_{examen.id}.xlsx'

    workbook.save(response)
    return response

from facturation.models import Facture

def factures_imagerie(request):
    factures = Facture.objects.filter(imagerie__isnull=False).select_related('patient', 'imagerie')
    return render(request, 'facturation/factures_par_imagerie.html', {
        'factures': factures
    })
