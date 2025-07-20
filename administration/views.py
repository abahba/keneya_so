from django.shortcuts import render, get_object_or_404, redirect
from .models import MembreAdministratif, HistoriqueAdministratif
from .forms import MembreAdministratifForm, MembreAdministratifFilterForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.utils import get_column_letter

@login_required
def liste_membres(request):
    """Affiche la liste des membres administratifs avec filtres"""
    membres = MembreAdministratif.objects.all().order_by('nom_complet')
    form = MembreAdministratifFilterForm(request.GET or None)

    if form.is_valid():
        nom = form.cleaned_data.get('nom')
        fonction = form.cleaned_data.get('fonction')
        actif = form.cleaned_data.get('actif')

        if nom:
            membres = membres.filter(nom_complet__icontains=nom)
        if fonction:
            membres = membres.filter(fonction__icontains=fonction)
        if actif == 'oui':
            membres = membres.filter(actif=True)
        elif actif == 'non':
            membres = membres.filter(actif=False)

    return render(request, 'administration/liste_membres.html', {
        'membres': membres,
        'form': form
    })

@login_required
def ajouter_membre(request):
    """Ajoute un nouveau membre administratif"""
    if request.method == 'POST':
        form = MembreAdministratifForm(request.POST)
        if form.is_valid():
            membre = form.save()
            
            # Enregistrement dans l'historique
            HistoriqueAdministratif.objects.create(
                membre=membre,
                utilisateur=request.user,
                action="Ajout",
                details=f"{membre.nom_complet} a été ajouté au personnel administratif"
            )
            
            messages.success(request, f"Le membre {membre.nom_complet} a été ajouté avec succès.")
            return redirect('administration:liste_membres')
    else:
        form = MembreAdministratifForm()
    
    return render(request, 'administration/ajouter_membre.html', {'form': form})

@login_required
def modifier_membre(request, pk):
    """Modifie un membre administratif existant"""
    membre = get_object_or_404(MembreAdministratif, pk=pk)
    
    if request.method == 'POST':
        form = MembreAdministratifForm(request.POST, instance=membre)
        if form.is_valid():
            membre = form.save()
            
            # Enregistrement dans l'historique
            HistoriqueAdministratif.objects.create(
                membre=membre,
                utilisateur=request.user,
                action="Modification",
                details=f"Modification des informations de {membre.nom_complet}"
            )
            
            messages.success(request, f"Les informations de {membre.nom_complet} ont été mises à jour.")
            return redirect('administration:liste_membres')
    else:
        form = MembreAdministratifForm(instance=membre)
    
    return render(request, 'administration/modifier_membre.html', {
        'form': form,
        'membre': membre
    })

@login_required
def supprimer_membre(request, pk):
    """Supprime un membre administratif"""
    membre = get_object_or_404(MembreAdministratif, pk=pk)
    
    if request.method == 'POST':
        # Enregistrement dans l'historique avant suppression
        HistoriqueAdministratif.objects.create(
            membre=membre,
            utilisateur=request.user,
            action="Suppression",
            details=f"{membre.nom_complet} a été supprimé du personnel administratif"
        )
        
        membre.delete()
        messages.success(request, f"Le membre {membre.nom_complet} a été supprimé.")
        return redirect('administration:liste_membres')
    
    return render(request, 'administration/supprimer_membre.html', {'membre': membre})

@login_required
def export_membres_pdf(request):
    """Exporte la liste des membres en PDF"""
    membres = MembreAdministratif.objects.all()
    template = get_template("administration/export_membres_pdf.html")
    html = template.render({'membres': membres})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="membres_administratifs.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

@login_required
def export_membres_excel(request):
    """Exporte la liste des membres en Excel"""
    membres = MembreAdministratif.objects.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Membres Administratifs"

    # En-têtes
    headers = ["Nom complet", "Fonction", "Téléphone", "Email", "Date embauche", "Statut"]
    sheet.append(headers)

    # Données
    for membre in membres:
        sheet.append([
            membre.nom_complet,
            membre.fonction,
            membre.telephone,
            membre.email,
            membre.date_embauche.strftime('%d/%m/%Y'),
            "Actif" if membre.actif else "Inactif"
        ])

    # Ajustement des colonnes
    for i, column in enumerate(sheet.columns, 1):
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(i)].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=membres_administratifs.xlsx'
    workbook.save(response)
    return response

@login_required
def historique_membres(request):
    """Affiche l'historique des actions administratives"""
    historiques = HistoriqueAdministratif.objects.select_related('membre', 'utilisateur').order_by('-date_action')
    return render(request, 'administration/historique.html', {'historiques': historiques})
